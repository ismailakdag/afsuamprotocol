"""
CSV Export Utilities for AFSUAM Measurement System.

Supports multiple export formats: CSV, JSON, Excel.
"""

import csv
import json
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

from protocols.base import ProtocolResult


class CSVExporter:
    """
    Exports protocol results to multiple formats.
    
    Supported formats:
    - CSV (default)
    - JSON
    - Excel (requires openpyxl)
    """
    
    # STEP record headers - comprehensive with all stats
    # STEP record headers - comprehensive with all stats
    STEP_HEADERS = [
        "record_type", "timestamp", "station", "ref_antenna_name", "repeat", "run_id",
        "beam_state", "port_config", "angle_deg", "v_ch1", "v_ch2", "dwell_s",
        "active_antennas", "tags_total",
        # Ant1 stats
        "ant1_unique_epc_n", "ant1_targets_seen_n", "ant1_nontarget_n",
        "ant1_total_reads", "ant1_rssi_min", "ant1_rssi_max", "ant1_rssi_avg",
        "ant1_missed_suffixes", "ant1_missed_labels",
        # Ant2 stats
        "ant2_unique_epc_n", "ant2_targets_seen_n", "ant2_nontarget_n",
        "ant2_total_reads", "ant2_rssi_min", "ant2_rssi_max", "ant2_rssi_avg",
        "ant2_missed_suffixes", "ant2_missed_labels"
    ]
    
    # TAGSTEP record headers - clean per-antenna data only
    TAGSTEP_HEADERS = [
        "record_type", "timestamp", "station", "ref_antenna_name", "repeat", "run_id",
        "beam_state", "port_config", "angle_deg", "dwell_s",
        "active_antennas",
        "tag_label", "tag_suffix", "tag_location",
        "ant1_seen", "ant1_rssi", "ant1_count", "ant1_phase",
        "ant2_seen", "ant2_rssi", "ant2_count", "ant2_phase"
    ]
    
    # UNION record headers - with best beam, RSSI, margin, and confidence per tag
    UNION_HEADERS = [
        "record_type", "timestamp", "station", "ref_antenna_name", "repeat", "run_id",
        "port_config", "dwell_s", "active_antennas", "tags_total",
        "ant1_unique_epcs", "ant2_unique_epcs",
        "ant1_targets_seen", "ant2_targets_seen",
        "ant1_missed_suffixes", "ant1_missed_labels",
        "ant2_missed_suffixes", "ant2_missed_labels",
        "ant1_best_beam_per_tag", "ant1_best_rssi_per_tag",
        "ant1_seen_beams_n_per_tag", "ant1_best_margin_per_tag",
        "ant1_tie_flag_per_tag", "ant1_best_confidence_per_tag",
        "ant2_best_beam_per_tag", "ant2_best_rssi_per_tag",
        "ant2_seen_beams_n_per_tag", "ant2_best_margin_per_tag",
        "ant2_tie_flag_per_tag", "ant2_best_confidence_per_tag"
    ]
    
    def __init__(self, output_dir: str = "outputs"):
        """
        Initialize exporter.
        
        Args:
            output_dir: Base directory for output files
        """
        self.base_output_dir = Path(output_dir)
        self.base_output_dir.mkdir(parents=True, exist_ok=True)
        # For backward compatibility
        self.output_dir = self.base_output_dir
    
    def _sanitize_name(self, name: str) -> str:
        """Sanitize name for use in filename."""
        # Replace spaces and special chars with underscores
        import re
        sanitized = re.sub(r'[^\w\-]', '_', name)
        # Remove multiple underscores
        sanitized = re.sub(r'_+', '_', sanitized)
        return sanitized.strip('_')
    
    def _get_date_folder(self) -> Path:
        """Get date-based output folder, creating if needed."""
        date_str = datetime.now().strftime("%Y-%m-%d")
        date_folder = self.base_output_dir / date_str
        date_folder.mkdir(parents=True, exist_ok=True)
        return date_folder
    
    def generate_filename(
        self,
        station_name: str,
        ref_antenna_name: str,
        mode: str = "LCR",
        format: str = "csv"
    ) -> Path:
        """
        Generate descriptive filename with date-based folder.
        
        Args:
            station_name: Station name (e.g., "AFSUAM Test-Bed")
            ref_antenna_name: Reference antenna name (e.g., "PhasedArray")
            mode: Protocol mode ("LCR" or "Inventory")
            format: File extension
        
        Returns:
            Full path to output file
        
        Example:
            outputs/2026-01-07/AFSUAM_TestBed_PhasedArray_LCR_130736.csv
        """
        date_folder = self._get_date_folder()
        
        # Sanitize names
        station = self._sanitize_name(station_name)
        ref_ant = self._sanitize_name(ref_antenna_name)
        
        # Generate timestamp
        timestamp = datetime.now().strftime("%H%M%S")
        
        # Build filename
        filename = f"{station}_{ref_ant}_{mode}_{timestamp}.{format}"
        
        return date_folder / filename
    
    def export_protocol_result(
        self,
        result: ProtocolResult,
        filename: Optional[str] = None,
        format: str = "csv",
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """
        Export protocol result to file.
        
        Args:
            result: ProtocolResult to export
            filename: Output filename (auto-generated if None)
            format: Export format ("csv", "json", "excel")
            metadata: Additional metadata to include
        
        Returns:
            Path to exported file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = "xlsx" if format == "excel" else format
            filename = f"protocol_{timestamp}.{ext}"
        
        filepath = self.output_dir / filename
        
        if format == "json":
            return self._export_json(result, filepath, metadata)
        elif format == "excel":
            return self._export_excel(result, filepath, metadata)
        else:
            return self._export_csv(result, filepath, metadata)
    
    def export_to_path(
        self,
        result: ProtocolResult,
        filepath: Path,
        metadata: Optional[Dict] = None
    ) -> str:
        """
        Export protocol result to specified absolute path.
        
        Unlike export_protocol_result, this method does NOT prepend output_dir.
        Use this when you already have the full path.
        
        Args:
            result: ProtocolResult to export
            filepath: Full path to output file (Path object)
            metadata: Additional metadata to include
        
        Returns:
            Path to exported file
        """
        # Ensure parent directory exists
        filepath = Path(filepath)
        filepath.parent.mkdir(parents=True, exist_ok=True)
        
        ext = filepath.suffix.lower()
        if ext == ".json":
            return self._export_json(result, filepath, metadata)
        elif ext in [".xlsx", ".xls"]:
            return self._export_excel(result, filepath, metadata)
        else:
            return self._export_csv(result, filepath, metadata)
    
    def _export_csv(
        self,
        result: ProtocolResult,
        filepath: Path,
        metadata: Optional[Dict] = None
    ) -> str:
        """Export to CSV format with comprehensive metadata."""
        
        # Build comprehensive RUN_METADATA
        run_metadata = {
            # Basic info
            "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "station": result.station_name,
            "ref_antenna": result.ref_antenna_name,
            "start_time": result.start_time,
            "end_time": result.end_time,
            "success": str(result.success),
            # Run identification
            "run_id": result.run_id,
            "data_valid": str(result.data_valid),
            "validation_errors": "|".join(result.validation_errors),
            # Protocol configuration
            "protocol_dwell_s": result.protocol_dwell_s,
            "protocol_repeats": result.protocol_repeats,
            "port_config": result.port_config,
            "beam_sequence": result.beam_sequence,
            "active_antennas": "|".join(str(a) for a in result.active_antennas),
            "tie_break_rule": result.tie_break_rule,
            # Target configuration
            "targets_configured_suffixes": "|".join(result.targets_configured_suffixes),
            "targets_configured_labels": "|".join(result.targets_configured_labels),
            # Hardware status
            "mcu_connected": "1" if result.mcu_connected else "0",
            "reader_connected": "1" if result.reader_connected else "0",
            "port2_enabled": "1" if result.port2_enabled else "0",
            # Antenna health
            "ant1_health": result.ant1_health,
            "ant2_health": result.ant2_health,
            "ant1_warning": result.ant1_warning,
            "ant2_warning": result.ant2_warning,
            # Notes
            "notes": result.notes,
            # Additional user metadata
            **(metadata or {})
        }
        
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            
            # Write metadata block
            writer.writerow(["# RUN_METADATA"])
            for key, value in run_metadata.items():
                writer.writerow([f"# {key}", value])
            writer.writerow([])
            
            # Write step results with full stats
            if result.step_results:
                writer.writerow(["# STEP_ROWS"])
                writer.writerow(self.STEP_HEADERS)
                for i, step in enumerate(result.step_results):
                    # Calculate repeat number
                    repeat_num = (i // 3) + 1 if len(result.step_results) >= 3 else 1
                    row = self._step_to_row(step, result, repeat_num)
                    writer.writerow(row)
                writer.writerow([])
            
            # Write tag step results - clean format
            if result.tag_step_results:
                writer.writerow(["# TAGSTEP_ROWS"])
                writer.writerow(self.TAGSTEP_HEADERS)
                for i, ts in enumerate(result.tag_step_results):
                    # Calculate repeat based on position
                    tags_per_step = len(result.tag_step_results) // max(len(result.step_results), 1)
                    repeat_num = (i // (tags_per_step * 3)) + 1 if tags_per_step > 0 else 1
                    row = self._tagstep_to_row(ts, result, repeat_num)
                    writer.writerow(row)
                writer.writerow([])
            
            # Write union results with best beam
            if result.union_results:
                writer.writerow(["# UNION_ROWS"])
                writer.writerow(self.UNION_HEADERS)
                for union in result.union_results:
                    row = self._union_to_row(union, result)
                    writer.writerow(row)
        
        print(f"Exported CSV: {filepath}")
        return str(filepath)
    
    def _export_json(
        self,
        result: ProtocolResult,
        filepath: Path,
        metadata: Optional[Dict] = None
    ) -> str:
        """Export to JSON format."""
        data = {
            "metadata": {
                "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "station": result.station_name,
                "ref_antenna": result.ref_antenna_name,
                "start_time": result.start_time,
                "end_time": result.end_time,
                "success": result.success,
                "protocol_dwell_s": result.protocol_dwell_s,
                "protocol_repeats": result.protocol_repeats,
                "port_config": result.port_config,
                "beam_sequence": result.beam_sequence,
                "active_antennas": result.active_antennas,
                "mcu_connected": result.mcu_connected,
                "reader_connected": result.reader_connected,
                "port2_enabled": result.port2_enabled,
                "ant1_health": result.ant1_health,
                "ant2_health": result.ant2_health,
                **(metadata or {})
            },
            "step_results": [],
            "tag_step_results": [],
            "union_results": []
        }
        
        # Convert step results
        for step in result.step_results:
            data["step_results"].append({
                "timestamp": step.timestamp,
                "beam_state": step.beam_state,
                "angle_deg": step.angle_deg,
                "port_config": step.port_config,
                "v_ch1": step.v_ch1,
                "v_ch2": step.v_ch2,
                "dwell_s": step.dwell_s,
                "active_antennas": step.active_antennas,
                "tags_total": step.tags_total,
                "ant1_unique_epc_n": step.ant1_unique_epc_n,
                "ant1_targets_seen_n": step.ant1_targets_seen_n,
                "ant1_total_reads": step.ant1_total_reads,
                "ant1_rssi_min": step.ant1_rssi_min,
                "ant1_rssi_max": step.ant1_rssi_max,
                "ant1_rssi_avg": step.ant1_rssi_avg,
                "ant2_unique_epc_n": step.ant2_unique_epc_n,
                "ant2_targets_seen_n": step.ant2_targets_seen_n,
                "ant2_total_reads": step.ant2_total_reads,
                "ant2_rssi_min": step.ant2_rssi_min,
                "ant2_rssi_max": step.ant2_rssi_max,
                "ant2_rssi_avg": step.ant2_rssi_avg
            })
        
        # Convert tag step results
        for ts in result.tag_step_results:
            data["tag_step_results"].append({
                "timestamp": ts.timestamp,
                "beam_state": ts.beam_state,
                "tag_label": ts.tag_label,
                "tag_suffix": ts.tag_suffix,
                "tag_location": ts.tag_location,
                "active_antennas": ts.active_antennas,
                "ant1_seen": ts.ant1_seen,
                "ant1_rssi": ts.ant1_rssi,
                "ant1_count": ts.ant1_count,
                "ant1_phase": ts.ant1_phase,
                "ant2_seen": ts.ant2_seen,
                "ant2_rssi": ts.ant2_rssi,
                "ant2_count": ts.ant2_count,
                "ant2_phase": ts.ant2_phase
            })
        
        # Convert union results
        for union in result.union_results:
            data["union_results"].append({
                "timestamp": union.timestamp,
                "repeat": union.repeat,
                "port_config": union.port_config,
                "dwell_s": union.dwell_s,
                "active_antennas": union.active_antennas,
                "tags_total": union.tags_total,
                "ant1_unique_epcs": union.ant1_unique_epcs,
                "ant2_unique_epcs": union.ant2_unique_epcs,
                "ant1_targets_seen": union.ant1_targets_seen,
                "ant2_targets_seen": union.ant2_targets_seen,
                "ant1_best_beam": union.ant1_best_beam,
                "ant2_best_beam": union.ant2_best_beam,
                "ant1_best_rssi": union.ant1_best_rssi,
                "ant2_best_rssi": union.ant2_best_rssi
            })
        
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        print(f"Exported JSON: {filepath}")
        return str(filepath)
    
    def _export_excel(
        self,
        result: ProtocolResult,
        filepath: Path,
        metadata: Optional[Dict] = None
    ) -> str:
        """Export to Excel format (requires openpyxl)."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to CSV if openpyxl not available
            print("openpyxl not available, falling back to CSV")
            csv_path = filepath.with_suffix(".csv")
            return self._export_csv(result, csv_path, metadata)
        
        wb = Workbook()
        
        # Metadata sheet
        ws_meta = wb.active
        ws_meta.title = "Metadata"
        ws_meta.append(["Key", "Value"])
        ws_meta.append(["Station", result.station_name])
        ws_meta.append(["Reference Antenna", result.ref_antenna_name])
        ws_meta.append(["Start Time", result.start_time])
        ws_meta.append(["End Time", result.end_time])
        ws_meta.append(["Success", str(result.success)])
        ws_meta.append(["Protocol Dwell", result.protocol_dwell_s])
        ws_meta.append(["Protocol Repeats", result.protocol_repeats])
        ws_meta.append(["Port Config", result.port_config])
        ws_meta.append(["Beam Sequence", result.beam_sequence])
        ws_meta.append(["Active Antennas", "|".join(str(a) for a in result.active_antennas)])
        ws_meta.append(["MCU Connected", result.mcu_connected])
        ws_meta.append(["Reader Connected", result.reader_connected])
        ws_meta.append(["Ant1 Health", result.ant1_health])
        ws_meta.append(["Ant2 Health", result.ant2_health])
        
        # Style header
        for cell in ws_meta[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DBEAFE", fill_type="solid")
        
        # Step Results sheet
        if result.step_results:
            ws_steps = wb.create_sheet("Step Results")
            ws_steps.append(self.STEP_HEADERS)
            for cell in ws_steps[1]:
                cell.font = Font(bold=True)
            for i, step in enumerate(result.step_results):
                repeat_num = (i // 3) + 1
                ws_steps.append(self._step_to_row(step, result, repeat_num))
        
        # Union Results sheet
        if result.union_results:
            ws_union = wb.create_sheet("Union Results")
            ws_union.append(self.UNION_HEADERS)
            for cell in ws_union[1]:
                cell.font = Font(bold=True)
            for union in result.union_results:
                ws_union.append(self._union_to_row(union, result))
        
        wb.save(filepath)
        print(f"Exported Excel: {filepath}")
        return str(filepath)
    
    def _step_to_row(self, step, result: ProtocolResult, repeat: int = 1) -> List:
        """Convert StepResult to row with all stats."""
        # Calculate nontarget counts
        ant1_nontarget = step.ant1_unique_epc_n - step.ant1_targets_seen_n
        ant2_nontarget = step.ant2_unique_epc_n - step.ant2_targets_seen_n
        
        return [
            "STEP",
            step.timestamp,
            result.station_name,
            result.ref_antenna_name,
            repeat,
            result.run_id,
            step.beam_state,
            step.port_config,
            step.angle_deg,
            step.v_ch1,
            step.v_ch2,
            step.dwell_s,
            "|".join(str(a) for a in step.active_antennas),
            step.tags_total,
            # Ant1
            step.ant1_unique_epc_n,
            step.ant1_targets_seen_n,
            ant1_nontarget,
            step.ant1_total_reads,
            f"{step.ant1_rssi_min:.1f}" if step.ant1_rssi_min is not None else "",
            f"{step.ant1_rssi_max:.1f}" if step.ant1_rssi_max is not None else "",
            f"{step.ant1_rssi_avg:.1f}" if step.ant1_rssi_avg is not None else "",
            "|".join(step.ant1_missed_suffixes),
            "|".join(step.ant1_missed_labels),
            # Ant2
            step.ant2_unique_epc_n,
            step.ant2_targets_seen_n,
            ant2_nontarget,
            step.ant2_total_reads,
            f"{step.ant2_rssi_min:.1f}" if step.ant2_rssi_min is not None else "",
            f"{step.ant2_rssi_max:.1f}" if step.ant2_rssi_max is not None else "",
            f"{step.ant2_rssi_avg:.1f}" if step.ant2_rssi_avg is not None else "",
            "|".join(step.ant2_missed_suffixes),
            "|".join(step.ant2_missed_labels)
        ]
    
    def _tagstep_to_row(self, ts, result: ProtocolResult, repeat: int = 1) -> List:
        """Convert TagStepResult to row - clean per-antenna format."""
        return [
            "TAGSTEP",
            ts.timestamp,
            result.station_name,
            result.ref_antenna_name,
            repeat,
            result.run_id,
            ts.beam_state,
            result.port_config,
            ts.angle_deg if hasattr(ts, 'angle_deg') else "",  # Support angle_deg if available
            result.protocol_dwell_s,
            "|".join(str(a) for a in ts.active_antennas),
            ts.tag_label,
            ts.tag_suffix,
            ts.tag_location,
            1 if ts.ant1_seen else 0,
            f"{ts.ant1_rssi:.1f}" if ts.ant1_rssi is not None else "",
            ts.ant1_count,
            f"{ts.ant1_phase:.1f}" if ts.ant1_phase is not None else "",
            1 if ts.ant2_seen else 0,
            f"{ts.ant2_rssi:.1f}" if ts.ant2_rssi is not None else "",
            ts.ant2_count,
            f"{ts.ant2_phase:.1f}" if ts.ant2_phase is not None else ""
        ]
    
    def _union_to_row(self, union, result: ProtocolResult) -> List:
        """Convert UnionResult to row with best beam per tag."""
        # Format best beam per tag: suffix:beam|suffix:beam|...
        ant1_best_beam_str = "|".join(f"{k}:{v}" for k, v in union.ant1_best_beam.items())
        ant2_best_beam_str = "|".join(f"{k}:{v}" for k, v in union.ant2_best_beam.items())
        
        # Format best RSSI per tag: suffix:rssi|suffix:rssi|...
        ant1_best_rssi_str = "|".join(f"{k}:{v:.1f}" for k, v in union.ant1_best_rssi.items())
        ant2_best_rssi_str = "|".join(f"{k}:{v:.1f}" for k, v in union.ant2_best_rssi.items())
        
        # Format seen beams count per tag: suffix:n|...
        ant1_seen_beams_str = "|".join(f"{k}:{v}" for k, v in union.ant1_seen_beams_n.items())
        ant2_seen_beams_str = "|".join(f"{k}:{v}" for k, v in union.ant2_seen_beams_n.items())
        
        # Format margin per tag: suffix:margin_db|... (NA for None)
        ant1_margin_str = "|".join(
            f"{k}:{v:.1f}" if v is not None else f"{k}:NA"
            for k, v in union.ant1_best_margin.items()
        )
        ant2_margin_str = "|".join(
            f"{k}:{v:.1f}" if v is not None else f"{k}:NA"
            for k, v in union.ant2_best_margin.items()
        )
        
        # Format tie flag per tag: suffix:0|suffix:1|...
        ant1_tie_str = "|".join(f"{k}:{v}" for k, v in union.ant1_tie_flag.items())
        ant2_tie_str = "|".join(f"{k}:{v}" for k, v in union.ant2_tie_flag.items())
        
        # Format confidence per tag: suffix:HIGH|suffix:MED|...
        ant1_conf_str = "|".join(f"{k}:{v}" for k, v in union.ant1_best_confidence.items())
        ant2_conf_str = "|".join(f"{k}:{v}" for k, v in union.ant2_best_confidence.items())
        
        return [
            "UNION",
            union.timestamp,
            result.station_name,
            result.ref_antenna_name,
            union.repeat,
            result.run_id,
            union.port_config,
            union.dwell_s,
            "|".join(str(a) for a in union.active_antennas),
            union.tags_total,
            union.ant1_unique_epcs,
            union.ant2_unique_epcs,
            union.ant1_targets_seen,
            union.ant2_targets_seen,
            "|".join(union.ant1_missed_suffixes) or "-",
            "|".join(union.ant1_missed_labels) or "-",
            "|".join(union.ant2_missed_suffixes) or "-",
            "|".join(union.ant2_missed_labels) or "-",
            ant1_best_beam_str,
            ant1_best_rssi_str,
            ant1_seen_beams_str,
            ant1_margin_str,
            ant1_tie_str,
            ant1_conf_str,
            ant2_best_beam_str,
            ant2_best_rssi_str,
            ant2_seen_beams_str,
            ant2_margin_str,
            ant2_tie_str,
            ant2_conf_str
        ]
    
    def export_live_snapshot(
        self,
        inventory: Dict[str, Dict],
        filename: Optional[str] = None,
        format: str = "csv",
        port_config: int = 0,
        angle: float = 0.0,
        v_ch1: float = 0.0,
        v_ch2: float = 0.0
    ) -> str:
        """
        Export live inventory snapshot.
        
        Args:
            inventory: Current inventory data
            filename: Output filename
            format: Export format
            port_config: Current port configuration
            angle: Current beam angle
            v_ch1, v_ch2: Current voltages
        
        Returns:
            Path to exported file
        """
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            ext = "xlsx" if format == "excel" else format
            filename = f"live_snapshot_{timestamp}.{ext}"
        
        filepath = self.output_dir / filename
        
        if format == "json":
            data = {
                "timestamp": datetime.now().isoformat(),
                "beam": {"port_config": port_config, "angle": angle, "v_ch1": v_ch1, "v_ch2": v_ch2},
                "tags": []
            }
            for epc, info in inventory.items():
                data["tags"].append({
                    "epc": epc,
                    "suffix": epc[-4:] if len(epc) >= 4 else epc,
                    "rssi": info.get("rssi", -99),
                    "phase": info.get("phase", 0),
                    "count": info.get("count", 0),
                    "antenna": info.get("antenna", 1)
                })
            
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
        else:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "timestamp", "epc", "suffix", "count", "rssi",
                    "phase", "doppler", "antenna",
                    "port_config", "angle_deg", "v_ch1", "v_ch2"
                ])
                
                for epc, info in inventory.items():
                    writer.writerow([
                        info.get("timestamp", ""),
                        epc,
                        epc[-4:] if len(epc) >= 4 else epc,
                        info.get("count", 0),
                        info.get("rssi", -99.0),
                        info.get("phase", 0.0),
                        info.get("doppler", 0.0),
                        info.get("antenna", 1),
                        port_config,
                        angle,
                        f"{v_ch1:.3f}",
                        f"{v_ch2:.3f}"
                    ])
        
        print(f"Exported snapshot: {filepath}")
        return str(filepath)
